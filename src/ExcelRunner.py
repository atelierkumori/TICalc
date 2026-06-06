import os
import sys
import subprocess
import platform
import openpyxl
from openpyxl.styles import Font
from ReadExcelParts import read_global_params, read_environment_data
from CalcCore import perform_calculation
from Utils import COL_ERROR, COL_OUTPUT_START, N_OUTPUT_COLS, ROW_START

OUTPUT_KEYS = [
    "WCT", "WBGT_Ono", "UTCI_approx", "PMV_7730", "PPD_7730", "SETstar",
    "theta_EATF_oet", "theta_TVF_oet", "theta_SERFL_oet", "theta_ERFS_oet",
    "theta_SEHF_oet", "theta_SECF_oet", "representative_temp", "ETU",
    "theta_EATF_etu", "theta_TVF_etu", "theta_SERFL_etu", "theta_ERFS_etu",
    "theta_SEHF_etu", "theta_SECF_etu",
]

assert len(OUTPUT_KEYS) == N_OUTPUT_COLS, (
    f"Developer error: OUTPUT_KEYS length ({len(OUTPUT_KEYS)}) "
    f"does not match N_OUTPUT_COLS ({N_OUTPUT_COLS})"
)

COL_OUTPUT_END = COL_OUTPUT_START + N_OUTPUT_COLS - 1


def _is_file_open_in_excel(file_path):
    abs_path = os.path.abspath(file_path)
    if platform.system() == "Darwin":
        try:
            result = subprocess.run(
                ["lsof", abs_path],
                capture_output=True, text=True
            )
            return "Microsoft" in result.stdout
        except Exception:
            return False
    elif platform.system() == "Windows":
        dir_name = os.path.dirname(abs_path)
        base_name = os.path.basename(abs_path)
        lock_file = os.path.join(dir_name, f"~${base_name}")
        return os.path.exists(lock_file)
    return False


def _close_file_in_excel(file_path):
    abs_path = os.path.abspath(file_path)
    wb_name = os.path.basename(abs_path)
    if platform.system() == "Darwin":
        script = f'''
tell application "Microsoft Excel"
    set wb_name to "{wb_name}"
    repeat with wb in workbooks
        if name of wb is wb_name then
            close wb saving yes
            exit repeat
        end if
    end repeat
end tell
'''
        subprocess.run(["osascript", "-e", script], capture_output=True)
    elif platform.system() == "Windows":
        try:
            import win32com.client
            excel = win32com.client.GetActiveObject("Excel.Application")
            for wb in excel.Workbooks:
                if wb.Name == wb_name:
                    wb.Close(SaveChanges=True)
                    break
        except Exception:
            pass


def _open_file_in_excel(file_path):
    abs_path = os.path.abspath(file_path)
    if platform.system() == "Darwin":
        subprocess.run(["open", "-a", "Microsoft Excel", abs_path])
    elif platform.system() == "Windows":
        os.startfile(abs_path)


def _ws_to_2d(ws, min_row, max_row, min_col, max_col):
    result = []
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row,
        min_col=min_col, max_col=max_col,
        values_only=True
    ):
        result.append(list(row))
    return result


def run_excel_calculation(file_path, sheet_name="Data"):
    abs_path = os.path.abspath(file_path)

    if _is_file_open_in_excel(abs_path):
        raise RuntimeError(
            f"'{os.path.basename(abs_path)}' is currently open in Excel.\n"
            "Please close the file in Excel and try again."
        )

    try:
        wb = openpyxl.load_workbook(abs_path)

        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Required sheet '{sheet_name}' not found.")

        ws = wb[sheet_name]

        # --- Loading of header range (A1:S7) ---
        header_data = _ws_to_2d(ws, 1, 7, 1, 19)

        global_params = read_global_params(header_data)

        # --- Detected the last row of data ---
        last_row = ROW_START
        for row in range(ws.max_row, ROW_START - 1, -1):
            if ws.cell(row=row, column=3).value is not None:
                last_row = row
                break

        if last_row < ROW_START:
            print("No data rows found.")
            return

        # --- Loading of data ---
        data_range = _ws_to_2d(ws, ROW_START, last_row, 1, 19)

        # header (rows 1-7) + gap (rows 8,9) + data (rows ROW_START onwards)
        gap_rows = [[None] * 19 for _ in range(ROW_START - len(header_data) - 1)]
        full_data = header_data + gap_rows + data_range

        # --- All calculations are performed ---
        error_rows = {}
        env_data_list = read_environment_data(full_data, global_params, error_rows)

        results_by_row = {}
        for record in env_data_list:
            row = record["row"]
            results = perform_calculation(global_params, record)
            if results:
                results_by_row[row] = [results.get(k, None) for k in OUTPUT_KEYS]

        # --- Construct a 2D array for output ---
        blank = [None] * N_OUTPUT_COLS
        all_output_values = []
        all_error_values = []

        for row in range(ROW_START, last_row + 1):
            if row in results_by_row:
                all_output_values.append(results_by_row[row])
                all_error_values.append(None)
            elif row in error_rows:
                all_output_values.append(blank[:])
                all_error_values.append(error_rows[row])
            else:
                all_output_values.append(blank[:])
                all_error_values.append(None)

        # --- Writing of output values ---
        arial10 = Font(name='Arial', size=10)

        for i, row_values in enumerate(all_output_values):
            excel_row = ROW_START + i
            for j, val in enumerate(row_values):
                cell = ws.cell(row=excel_row, column=COL_OUTPUT_START + j)
                cell.value = val
                cell.font = arial10
                cell.number_format = '0.00'

        for i, err_val in enumerate(all_error_values):
            excel_row = ROW_START + i
            cell = ws.cell(row=excel_row, column=COL_ERROR)
            cell.value = err_val
            cell.font = arial10


        wb.save(abs_path)
        wb.close()
        print("Calculation completed and saved.")

    except Exception:
        raise